import openpyxl as oxl
import os
import json


def read_json(json_file_path: str):
    # Check if the file exists and is not empty
    if not os.path.isfile(json_file_path) or os.path.getsize(json_file_path) == 0:
        print(f"The file {json_file_path} does not exist or is empty.")
        return {}
    
    # Reading JSON file content into a string
    with open(json_file_path, 'r') as json_file:
        json_string = json_file.read()

    json_string = json_string.replace(":,", ": null,")

    # Reading JSON data from a file
    with open(json_file_path, 'r') as json_file:
        loaded_data = json.loads(json_string)

    return loaded_data


def read_excel(excel_file_path: str) -> dict:
    # Check if the file exists and is not empty
    if not os.path.isfile(excel_file_path) or os.path.getsize(excel_file_path) == 0:
        print(f"The file {excel_file_path} does not exist or is empty.")
        return {}
    
    wb = oxl.load_workbook(excel_file_path)
    sheet_content = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        loaded_excel = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
            loaded_excel.append(list(row))

        loaded_columns = []
        loaded_dataset = [{} for _ in range(len(loaded_excel) - 1)]

        for col in range(len(loaded_excel[0])):
            if loaded_excel[0][col] is None:
                continue
            
            loaded_columns.append(loaded_excel[0][col])
            for row in range(len(loaded_excel) - 1):
                loaded_dataset[row][loaded_excel[0][col]] = loaded_excel[row + 1][col]
            
        sheet_content[sheetname] = {
            "dataset": loaded_dataset,
            "data_columns": loaded_columns,
            "data_size": len(loaded_dataset)
        }
    
    return sheet_content


def update_json_dict(target_dict: dict, path: list[str], value: any):
    """
    Update a specific value in a nested dictionary.

    Parameters:
    target_dict (dict): The dictionary to be updated.
    path_to_value (list[str]): The path to the value to be updated. Each element in the list represents a key in the dictionary.
    new_value (any): The new value to be set.

    Returns:
    dict: The updated dictionary.
    """

    if path is None:
        return target_dict

    if len(path) == 0:  # the last position
        return value

    target_dict[path[0]] = update_json_dict(target_dict[path[0]], path[1:], value)

    return target_dict


def write_json(json_obj: dict | list, json_output_path: str, indent: int = 4, decoder: bool = True):
    dir = os.path.dirname(json_output_path)
    if not os.path.exists(dir):
        os.makedirs(dir)
        
    with open(json_output_path, 'w', encoding="utf-8") as json_file:
        json.dump(json_obj, json_file, indent=indent, ensure_ascii=not decoder)


if __name__ == "__main__":
    # Read
    json_r = read_json("../example/002.json")
    excel_r, columns, num = read_excel("../example/test.xlsx")

    print(json_r)
    print(excel_r)
    print(columns)
    print(num)

    # Write
    write_json(json_r, "../example/result/003.json")
